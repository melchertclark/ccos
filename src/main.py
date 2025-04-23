import os
import sys
import time
import random
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from pathlib import Path
import asyncio
import httpx
from openpyxl import load_workbook
from readability import Document
from bs4 import BeautifulSoup
import argparse

class URLScraper:
    def __init__(self):
        self.progress_window = None
        self.progress_var = None
        self.cancelled = False
        self.total_urls = 0
        self.processed_urls = 0

    def create_progress_window(self):
        self.progress_window = tk.Toplevel()
        self.progress_window.title("URL Scraper Progress")
        self.progress_window.geometry("400x150")
        
        self.progress_var = tk.DoubleVar()
        progress_bar = tk.Progressbar(
            self.progress_window, 
            variable=self.progress_var, 
            maximum=100
        )
        progress_bar.pack(pady=20, padx=20, fill=tk.X)
        
        self.status_label = tk.Label(self.progress_window, text="0 / 0 URLs processed")
        self.status_label.pack(pady=10)
        
        cancel_button = tk.Button(
            self.progress_window, 
            text="Cancel", 
            command=self.cancel_scraping
        )
        cancel_button.pack(pady=10)
        
        self.progress_window.protocol("WM_DELETE_WINDOW", self.cancel_scraping)
        self.progress_window.update()

    def cancel_scraping(self):
        self.cancelled = True
        if self.progress_window:
            self.progress_window.destroy()

    def update_progress(self):
        if self.progress_window and not self.cancelled:
            progress = (self.processed_urls / self.total_urls) * 100
            self.progress_var.set(progress)
            self.status_label.config(text=f"{self.processed_urls} / {self.total_urls} URLs processed")
            self.progress_window.update()

    def get_plain_text(self, html_content):
        try:
            doc = Document(html_content)
            return doc.summary()
        except Exception:
            soup = BeautifulSoup(html_content, 'html.parser')
            return soup.get_text(separator='\n', strip=True)

    async def scrape_url(self, client, label, url):
        try:
            response = await client.get(url, timeout=30.0)
            response.raise_for_status()
            return label, url, self.get_plain_text(response.text), None
        except Exception as e:
            return label, url, None, str(e)

    async def process_urls(self, urls, concurrency=1):
        async with httpx.AsyncClient() as client:
            tasks = []
            for label, url in urls:
                if self.cancelled:
                    break
                tasks.append(self.scrape_url(client, label, url))
                if len(tasks) >= concurrency:
                    results = await asyncio.gather(*tasks)
                    for result in results:
                        self.processed_urls += 1
                        self.update_progress()
                        yield result
                    tasks = []
            
            if tasks and not self.cancelled:
                results = await asyncio.gather(*tasks)
                for result in results:
                    self.processed_urls += 1
                    self.update_progress()
                    yield result

    def run(self, concurrency=1):
        root = tk.Tk()
        root.withdraw()

        file_path = filedialog.askopenfilename(
            title="Select Excel Workbook",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not file_path:
            return

        try:
            workbook = load_workbook(file_path)
            if "URL Whiteboard" not in workbook.sheetnames:
                messagebox.showerror("Error", "Required sheet 'URL Whiteboard' not found")
                return

            sheet = workbook["URL Whiteboard"]
            urls = []
            for row in sheet.iter_rows(min_row=1, max_col=2):
                if not row[1].value:  # Empty URL
                    break
                urls.append((row[0].value, row[1].value))

            if not urls:
                messagebox.showerror("Error", "No URLs found in the workbook")
                return

            self.total_urls = len(urls)
            self.create_progress_window()

            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
            output_dir = Path(file_path).parent
            output_file = output_dir / f"combined_{timestamp}.txt"
            log_file = output_dir / f"scrape_log_{timestamp}.txt"

            with open(output_file, 'w', encoding='utf-8') as out_f, \
                 open(log_file, 'w', encoding='utf-8') as log_f:
                
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                
                for label, url, content, error in loop.run_until_complete(
                    self.process_urls(urls, concurrency)
                ):
                    if error:
                        log_f.write(f"{label} – {url} – {error}\n")
                    else:
                        out_f.write(f"\n----- {label} -----\n\n")
                        out_f.write(content)
                        out_f.write("\n\n")

                if self.cancelled:
                    log_f.write(f"Run cancelled by user after {self.processed_urls} of {self.total_urls} URLs\n")

            if self.progress_window and not self.cancelled:
                self.progress_window.destroy()
                messagebox.showinfo(
                    "Complete",
                    f"Scraping completed!\nOutput files:\n{output_file}\n{log_file}"
                )

            # Open the output directory
            if sys.platform == "darwin":  # macOS
                os.system(f"open {output_dir}")
            else:  # Windows
                os.system(f"explorer {output_dir}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

def main():
    parser = argparse.ArgumentParser(description="URL Whiteboard Scraper")
    parser.add_argument(
        "--concurrency",
        type=int,
        default=1,
        help="Number of concurrent requests (default: 1)"
    )
    args = parser.parse_args()

    scraper = URLScraper()
    scraper.run(args.concurrency)

if __name__ == "__main__":
    main() 